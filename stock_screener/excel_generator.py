"""
Excel 報表產生器
產出精美格式化的 .xlsx 選股結果，含條件式格式、自動欄寬等
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import config


def generate_excel(results, total_scanned, output_path="report.xlsx"):
    """
    產生格式化 Excel 報表。

    Parameters
    ----------
    results : list[dict]  每個 dict 含:
        code, name, market, close, change_pct, volume_lots,
        daily_buy_reason, daily_trend_code, daily_trend_text,
        weekly_reason, weekly_trend_code, weekly_trend_text
    total_scanned : int
    output_path : str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "選股結果"

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_str = datetime.now().strftime("%Y-%m-%d")

    # ── 色彩定義 ──
    GOLD = "F0B90B"
    DARK_BG = "1E222D"
    HEADER_BG = "262A36"
    GREEN = "26A69A"
    RED = "EF5350"
    BLUE = "42A5F5"
    PURPLE = "CE93D8"
    WHITE = "E0E0E0"
    LIGHT_GRAY = "888888"

    header_font = Font(name="微軟正黑體", size=11, bold=True, color=GOLD)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    data_font = Font(name="微軟正黑體", size=10, color=WHITE)
    data_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    green_font = Font(name="微軟正黑體", size=10, bold=True, color=GREEN)
    red_font = Font(name="微軟正黑體", size=10, bold=True, color=RED)
    thin_border = Border(
        bottom=Side(style="thin", color="333333")
    )

    # ── 標題區 ──
    title_fill = PatternFill(start_color="12141C", end_color="12141C", fill_type="solid")
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 箱波均戰法 — 選股結果　　{now}"
    title_cell.font = Font(name="微軟正黑體", size=16, bold=True, color=GOLD)
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:K2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"篩選條件：週K已買進 + 日K當日新買進訊號　|　"
        f"成交量 ≥ {config.MIN_VOLUME_LOTS} 張　|　股價 ≤ {config.MAX_PRICE} 元　|　"
        f"掃描 {total_scanned} 檔 → 符合 {len(results)} 檔"
    )
    subtitle_cell.font = Font(name="微軟正黑體", size=10, color=LIGHT_GRAY)
    subtitle_cell.fill = title_fill
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── 表頭 ──
    headers = [
        "#", "代碼", "名稱", "市場", "收盤價",
        "漲跌幅(%)", "成交量(張)", "日K買進原因",
        "日K均線狀態", "週K買進原因", "週K均線狀態"
    ]
    col_widths = [5, 10, 14, 8, 12, 12, 14, 14, 14, 14, 14]

    header_row = 4
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 28

    # ── 資料列 ──
    for i, r in enumerate(results, 1):
        row_num = header_row + i
        chg = r.get("change_pct", 0)
        price_font = green_font if chg >= 0 else red_font

        row_data = [
            i,
            r["code"],
            r["name"],
            r["market"],
            r["close"],
            round(chg, 2),
            r["volume_lots"],
            r["daily_buy_reason"],
            r["daily_trend_text"],
            r["weekly_reason"],
            r["weekly_trend_text"],
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = data_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 預設字型
            cell.font = data_font

            # 收盤價 & 漲跌幅 用漲跌色
            if col_idx == 5:  # 收盤價
                cell.font = price_font
                cell.number_format = "#,##0.00"
            elif col_idx == 6:  # 漲跌幅
                cell.font = price_font
                cell.number_format = "+0.00%;-0.00%"
                cell.value = chg / 100  # 轉為小數讓 Excel 格式化
            elif col_idx == 7:  # 成交量
                cell.number_format = "#,##0"
            elif col_idx == 9:  # 日K趨勢
                cell.font = _trend_font(r["daily_trend_code"])
            elif col_idx == 11:  # 週K趨勢
                cell.font = _trend_font(r["weekly_trend_code"])

        ws.row_dimensions[row_num].height = 22

    # ── 如果沒有結果 ──
    if not results:
        ws.merge_cells(f"A{header_row + 1}:K{header_row + 1}")
        no_data = ws.cell(row=header_row + 1, column=1, value="🔍 今日無符合條件的股票")
        no_data.font = Font(name="微軟正黑體", size=14, color=LIGHT_GRAY)
        no_data.fill = data_fill
        no_data.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row + 1].height = 50

    # ── 凍結窗格（凍結表頭） ──
    ws.freeze_panes = f"A{header_row + 1}"

    # ── 設定列印範圍 & 頁面設定 ──
    ws.sheet_properties.tabColor = GOLD

    # ── 儲存 ──
    abs_path = os.path.abspath(output_path)
    wb.save(abs_path)
    return abs_path


def _trend_font(trend_code):
    """根據趨勢代碼回傳字型顏色"""
    GREEN = "26A69A"
    RED = "EF5350"
    ORANGE = "FF9800"
    GRAY = "888888"

    if trend_code in (1, 2):
        color = GREEN
    elif trend_code in (-1, -2):
        color = RED
    elif trend_code == 3:
        color = ORANGE
    else:
        color = GRAY

    return Font(name="微軟正黑體", size=10, bold=True, color=color)
